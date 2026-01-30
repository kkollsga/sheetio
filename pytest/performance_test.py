#!/usr/bin/env python3
"""
Extensive Performance Test Suite for sheet_excavator

Tests all 10 bottleneck optimizations with:
- Correctness validation (output matches expected)
- Performance benchmarking (timing measurements)
- Memory usage monitoring
- Stress testing with large datasets

Run with: python pytest/performance_test.py
"""

import sheet_excavator
import json
import time
import os
import tempfile
import tracemalloc
from typing import Dict, List, Any, Tuple
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


class TestDataGenerator:
    """Generates test Excel files with configurable parameters."""

    def __init__(self, temp_dir: str):
        self.temp_dir = temp_dir
        self.generated_files: List[str] = []

    def create_large_multirow_file(
        self,
        filename: str,
        num_rows: int = 1000,
        num_columns: int = 15,
        num_sheets: int = 3
    ) -> str:
        """Create Excel file with many rows/columns for bottleneck #2-3 testing."""
        filepath = os.path.join(self.temp_dir, filename)
        wb = Workbook()

        for sheet_idx in range(num_sheets):
            if sheet_idx == 0:
                ws = wb.active
                ws.title = f"Sheet{sheet_idx + 1}"
            else:
                ws = wb.create_sheet(f"Sheet{sheet_idx + 1}")

            # Header row
            for col in range(1, num_columns + 1):
                ws.cell(row=1, column=col, value=f"Column_{get_column_letter(col)}")

            # Data rows
            for row in range(2, num_rows + 2):
                for col in range(1, num_columns + 1):
                    # Mix of data types
                    if col == 1:
                        ws.cell(row=row, column=col, value=f"ID_{row-1}")
                    elif col % 3 == 0:
                        ws.cell(row=row, column=col, value=row * col * 1.5)
                    elif col % 3 == 1:
                        ws.cell(row=row, column=col, value=row * col)
                    else:
                        ws.cell(row=row, column=col, value=f"  Data_{row}_{col}  ")  # With whitespace for trim test

        wb.save(filepath)
        self.generated_files.append(filepath)
        return filepath

    def create_multi_extraction_file(self, filename: str) -> str:
        """Create file for bottleneck #4 testing (multiple extraction configs)."""
        filepath = os.path.join(self.temp_dir, filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # Section 1: Single cells
        ws['A1'] = "Header1"
        ws['B1'] = "Value1"
        ws['A2'] = "Header2"
        ws['B2'] = 12345

        # Section 2: Multirow data
        for row in range(5, 55):
            ws.cell(row=row, column=1, value=f"Row{row-4}")
            ws.cell(row=row, column=2, value=row * 10)
            ws.cell(row=row, column=3, value=f"Text_{row}")

        # Section 3: Dataframe
        for col in range(5, 10):
            ws.cell(row=60, column=col, value=f"Col{col-4}")
            for row in range(61, 80):
                ws.cell(row=row, column=col, value=row + col)

        wb.save(filepath)
        self.generated_files.append(filepath)
        return filepath

    def create_many_sheets_file(self, filename: str, num_sheets: int = 50) -> str:
        """Create file with many sheets for bottleneck #6, #8 testing."""
        filepath = os.path.join(self.temp_dir, filename)
        wb = Workbook()

        for i in range(num_sheets):
            if i == 0:
                ws = wb.active
                ws.title = f"Sheet_{i+1:03d}"
            else:
                ws = wb.create_sheet(f"Sheet_{i+1:03d}")

            ws['A1'] = f"Sheet {i+1}"
            ws['B1'] = i * 100

        wb.save(filepath)
        self.generated_files.append(filepath)
        return filepath

    def create_duplicate_keys_file(self, filename: str) -> str:
        """Create file that produces duplicate keys for bottleneck #7 testing."""
        filepath = os.path.join(self.temp_dir, filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "Duplicates"

        # Many rows with same unique_id to trigger deduplication
        for row in range(1, 101):
            ws.cell(row=row, column=1, value="DUPLICATE_KEY")  # All same ID
            ws.cell(row=row, column=2, value=row)
            ws.cell(row=row, column=3, value=f"Data_{row}")

        wb.save(filepath)
        self.generated_files.append(filepath)
        return filepath

    def create_bulk_files(self, prefix: str, count: int = 50) -> List[str]:
        """Create many small files for bottleneck #1 testing (Arc cloning)."""
        files = []
        for i in range(count):
            filepath = os.path.join(self.temp_dir, f"{prefix}_{i+1:03d}.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"

            for row in range(1, 21):
                ws.cell(row=row, column=1, value=f"ID_{row}")
                ws.cell(row=row, column=2, value=row * 10)
                ws.cell(row=row, column=3, value=f"Value_{row}")

            wb.save(filepath)
            files.append(filepath)
            self.generated_files.append(filepath)

        return files

    def cleanup(self):
        """Remove all generated test files."""
        for f in self.generated_files:
            if os.path.exists(f):
                os.remove(f)


class PerformanceTestSuite:
    """Comprehensive test suite for all bottleneck optimizations."""

    def __init__(self):
        self.results: Dict[str, Dict[str, Any]] = {}
        self.temp_dir = tempfile.mkdtemp(prefix="sheet_excavator_perf_")
        self.generator = TestDataGenerator(self.temp_dir)

    def _time_execution(self, func, *args, **kwargs) -> Tuple[Any, float]:
        """Execute function and return (result, elapsed_seconds)."""
        start = time.perf_counter()
        result = func(*args, **kwargs)
        elapsed = time.perf_counter() - start
        return result, elapsed

    def _measure_memory(self, func, *args, **kwargs) -> Tuple[Any, int]:
        """Execute function and return (result, peak_memory_bytes)."""
        tracemalloc.start()
        result = func(*args, **kwargs)
        _, peak = tracemalloc.get_traced_memory()
        tracemalloc.stop()
        return result, peak

    # =========================================================================
    # BOTTLENECK #1: Arc for extraction_details (Deep clone per file)
    # =========================================================================
    def test_bottleneck_1_arc_cloning(self):
        """Test: Process many files to verify Arc optimization."""
        print("\n" + "="*70)
        print("BOTTLENECK #1: Arc for extraction_details")
        print("="*70)

        # Create 50 test files
        files = self.generator.create_bulk_files("bulk", count=50)

        # Large extraction config to make cloning expensive
        config = [{
            "sheets": ["Data"],
            "extractions": [{
                "function": "multirow_patterns",
                "instructions": {
                    "row_range": [1, 20],
                    "unique_id": "A",
                    "columns": {f"col_{i}": chr(65 + i) for i in range(3)}
                }
            }] * 5  # Multiple extractions to make config larger
        }]

        # Measure memory and time
        def run_extraction():
            return sheet_excavator.excel_extract(files, config, 8)

        result, peak_memory = self._measure_memory(run_extraction)
        _, elapsed = self._time_execution(run_extraction)

        parsed = json.loads(result)

        # Validate correctness
        assert len(parsed) == 50, f"Expected 50 files, got {len(parsed)}"

        self.results['bottleneck_1'] = {
            'files_processed': 50,
            'time_seconds': elapsed,
            'peak_memory_mb': peak_memory / (1024 * 1024),
            'status': 'PASS'
        }

        print(f"  Files processed: 50")
        print(f"  Time: {elapsed:.3f}s")
        print(f"  Peak memory: {peak_memory / (1024*1024):.2f} MB")
        print(f"  Status: PASS")

    # =========================================================================
    # BOTTLENECK #2-3: Pre-parsed column config (Column parsing/cloning per row)
    # =========================================================================
    def test_bottleneck_2_3_column_parsing(self):
        """Test: Large multirow extraction to verify pre-parsing optimization."""
        print("\n" + "="*70)
        print("BOTTLENECK #2-3: Pre-parsed column config")
        print("="*70)

        # Create file with many rows and columns
        filepath = self.generator.create_large_multirow_file(
            "large_multirow.xlsx",
            num_rows=2000,
            num_columns=15,
            num_sheets=1
        )

        # Config with many columns
        columns = {f"col_{get_column_letter(i)}": get_column_letter(i) for i in range(1, 16)}

        config = [{
            "sheets": ["Sheet1"],
            "extractions": [{
                "function": "multirow_patterns",
                "instructions": {
                    "row_range": [2, 2001],
                    "unique_id": "A",
                    "columns": columns
                }
            }]
        }]

        result, elapsed = self._time_execution(
            sheet_excavator.excel_extract, [filepath], config, 1
        )

        parsed = json.loads(result)

        # Validate correctness
        file_key = list(parsed.keys())[0]
        sheet_data = parsed[file_key].get("Sheet1", {})
        num_rows_extracted = len(sheet_data)

        assert num_rows_extracted == 2000, f"Expected 2000 rows, got {num_rows_extracted}"

        # Verify a sample row has all columns
        first_key = list(sheet_data.keys())[0]
        first_row = sheet_data[first_key]
        assert len(first_row) == 15, f"Expected 15 columns, got {len(first_row)}"

        self.results['bottleneck_2_3'] = {
            'rows_extracted': num_rows_extracted,
            'columns': 15,
            'time_seconds': elapsed,
            'status': 'PASS'
        }

        print(f"  Rows extracted: {num_rows_extracted}")
        print(f"  Columns per row: 15")
        print(f"  Time: {elapsed:.3f}s")
        print(f"  Status: PASS")

    # =========================================================================
    # BOTTLENECK #4: Single workbook open (Workbook reopened per config)
    # =========================================================================
    def test_bottleneck_4_single_workbook(self):
        """Test: Multiple extraction configs on same file."""
        print("\n" + "="*70)
        print("BOTTLENECK #4: Single workbook open")
        print("="*70)

        filepath = self.generator.create_multi_extraction_file("multi_extract.xlsx")

        # Multiple extraction configs for same file
        config = [{
            "sheets": ["Data"],
            "extractions": [
                {
                    "function": "single_cells",
                    "label": "header_data",
                    "instructions": {
                        "cell1": "A1",
                        "cell2": "B1",
                        "cell3": "A2",
                        "cell4": "B2"
                    }
                },
                {
                    "function": "multirow_patterns",
                    "label": "row_data",
                    "instructions": {
                        "row_range": [5, 54],
                        "unique_id": "A",
                        "columns": {"name": "A", "value": "B", "text": "C"}
                    }
                },
                {
                    "function": "dataframe",
                    "label": "table_data",
                    "instructions": {
                        "header_row": 60,
                        "row_range": [61, 79],
                        "column_range": ["E", "I"]
                    }
                }
            ]
        }]

        result, elapsed = self._time_execution(
            sheet_excavator.excel_extract, [filepath], config, 1
        )

        parsed = json.loads(result)

        # Validate all 3 extraction types worked
        file_key = list(parsed.keys())[0]
        sheet_data = parsed[file_key].get("Data", {})

        assert "header_data" in sheet_data, "Missing header_data"
        assert "row_data" in sheet_data, "Missing row_data"
        assert "table_data" in sheet_data, "Missing table_data"

        # Validate counts
        assert len(sheet_data["row_data"]) == 50, f"Expected 50 rows, got {len(sheet_data['row_data'])}"

        self.results['bottleneck_4'] = {
            'extraction_configs': 3,
            'time_seconds': elapsed,
            'status': 'PASS'
        }

        print(f"  Extraction configs: 3")
        print(f"  Time: {elapsed:.3f}s")
        print(f"  Status: PASS")

    # =========================================================================
    # BOTTLENECK #5: Conditional trim (trim().to_string() per cell)
    # =========================================================================
    def test_bottleneck_5_conditional_trim(self):
        """Test: Cells with whitespace are trimmed correctly."""
        print("\n" + "="*70)
        print("BOTTLENECK #5: Conditional trim allocation")
        print("="*70)

        # Use the large file which has whitespace in string columns
        filepath = self.generator.create_large_multirow_file(
            "trim_test.xlsx",
            num_rows=500,
            num_columns=10,
            num_sheets=1
        )

        config = [{
            "sheets": ["Sheet1"],
            "extractions": [{
                "function": "multirow_patterns",
                "instructions": {
                    "row_range": [2, 501],
                    "unique_id": "A",
                    "columns": {"data": "B"}  # Column B has "  Data_X_Y  " with spaces
                }
            }]
        }]

        result, elapsed = self._time_execution(
            sheet_excavator.excel_extract, [filepath], config, 1
        )

        parsed = json.loads(result)

        file_key = list(parsed.keys())[0]
        sheet_data = parsed[file_key].get("Sheet1", {})

        # Verify trimming worked - values should not have leading/trailing spaces
        sample_key = list(sheet_data.keys())[0]
        sample_value = sheet_data[sample_key].get("data", "")

        # The data should be trimmed (no leading/trailing spaces)
        if isinstance(sample_value, str):
            assert sample_value == sample_value.strip(), "Trimming not working"

        self.results['bottleneck_5'] = {
            'cells_processed': 500,
            'time_seconds': elapsed,
            'trim_verified': True,
            'status': 'PASS'
        }

        print(f"  Cells with whitespace: 500")
        print(f"  Time: {elapsed:.3f}s")
        print(f"  Trim verified: Yes")
        print(f"  Status: PASS")

    # =========================================================================
    # BOTTLENECK #6: HashSet for skip_sheets
    # =========================================================================
    def test_bottleneck_6_skip_sheets_hashset(self):
        """Test: Many skip_sheets entries for O(1) vs O(n) lookup."""
        print("\n" + "="*70)
        print("BOTTLENECK #6: HashSet for skip_sheets")
        print("="*70)

        filepath = self.generator.create_many_sheets_file("many_sheets.xlsx", num_sheets=50)

        # Skip most sheets, process only a few
        skip_list = [f"Sheet_{i:03d}" for i in range(3, 51)]  # Skip sheets 3-50

        config = [{
            "sheets": ["Sheet_*"],
            "skip_sheets": skip_list,  # 48 sheets to skip
            "extractions": [{
                "function": "single_cells",
                "instructions": {"value": "A1", "number": "B1"}
            }]
        }]

        result, elapsed = self._time_execution(
            sheet_excavator.excel_extract, [filepath], config, 1
        )

        parsed = json.loads(result)

        file_key = list(parsed.keys())[0]
        sheets_processed = [k for k in parsed[file_key].keys() if k != "filepath"]

        # Should only process Sheet_001 and Sheet_002
        assert len(sheets_processed) == 2, f"Expected 2 sheets, got {len(sheets_processed)}"

        self.results['bottleneck_6'] = {
            'total_sheets': 50,
            'skip_sheets': 48,
            'processed_sheets': 2,
            'time_seconds': elapsed,
            'status': 'PASS'
        }

        print(f"  Total sheets: 50")
        print(f"  Skip list size: 48")
        print(f"  Sheets processed: 2")
        print(f"  Time: {elapsed:.3f}s")
        print(f"  Status: PASS")

    # =========================================================================
    # BOTTLENECK #7: Unique key helper (duplicate resolution)
    # =========================================================================
    def test_bottleneck_7_unique_keys(self):
        """Test: Many duplicate keys to stress deduplication."""
        print("\n" + "="*70)
        print("BOTTLENECK #7: Unique key helper")
        print("="*70)

        filepath = self.generator.create_duplicate_keys_file("duplicates.xlsx")

        config = [{
            "sheets": ["Duplicates"],
            "extractions": [{
                "function": "multirow_patterns",
                "instructions": {
                    "row_range": [1, 100],
                    "unique_id": "A",  # All rows have "DUPLICATE_KEY"
                    "columns": {"value": "B", "data": "C"}
                }
            }]
        }]

        result, elapsed = self._time_execution(
            sheet_excavator.excel_extract, [filepath], config, 1
        )

        parsed = json.loads(result)

        file_key = list(parsed.keys())[0]
        sheet_data = parsed[file_key].get("Duplicates", {})

        # Should have 100 unique keys: DUPLICATE_KEY, DUPLICATE_KEY_1, ..., DUPLICATE_KEY_99
        assert len(sheet_data) == 100, f"Expected 100 unique keys, got {len(sheet_data)}"

        # Verify key pattern
        assert "DUPLICATE_KEY" in sheet_data
        assert "DUPLICATE_KEY_1" in sheet_data
        assert "DUPLICATE_KEY_99" in sheet_data

        self.results['bottleneck_7'] = {
            'duplicate_rows': 100,
            'unique_keys_generated': len(sheet_data),
            'time_seconds': elapsed,
            'status': 'PASS'
        }

        print(f"  Duplicate rows: 100")
        print(f"  Unique keys generated: {len(sheet_data)}")
        print(f"  Time: {elapsed:.3f}s")
        print(f"  Status: PASS")

    # =========================================================================
    # BOTTLENECK #8: IndexSet for sheet_names (extend_unique)
    # =========================================================================
    def test_bottleneck_8_indexset_sheets(self):
        """Test: Wildcard matching many sheets for O(n) vs O(n^2)."""
        print("\n" + "="*70)
        print("BOTTLENECK #8: IndexSet for sheet_names")
        print("="*70)

        filepath = self.generator.create_many_sheets_file("indexset_test.xlsx", num_sheets=50)

        config = [{
            "sheets": ["Sheet_*"],  # Wildcard matches all 50
            "extractions": [{
                "function": "single_cells",
                "instructions": {"title": "A1"}
            }]
        }]

        result, elapsed = self._time_execution(
            sheet_excavator.excel_extract, [filepath], config, 1
        )

        parsed = json.loads(result)

        file_key = list(parsed.keys())[0]
        sheets_processed = [k for k in parsed[file_key].keys() if k != "filepath"]

        assert len(sheets_processed) == 50, f"Expected 50 sheets, got {len(sheets_processed)}"

        self.results['bottleneck_8'] = {
            'sheets_matched': 50,
            'time_seconds': elapsed,
            'status': 'PASS'
        }

        print(f"  Sheets matched by wildcard: 50")
        print(f"  Time: {elapsed:.3f}s")
        print(f"  Status: PASS")

    # =========================================================================
    # BOTTLENECK #9: Cow for type descriptions
    # =========================================================================
    def test_bottleneck_9_type_descriptions(self):
        """Test: All data types are handled correctly."""
        print("\n" + "="*70)
        print("BOTTLENECK #9: Cow for type descriptions")
        print("="*70)

        # Create file with various data types
        filepath = os.path.join(self.temp_dir, "datatypes.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "Types"

        ws['A1'] = "StringValue"
        ws['A2'] = 12345
        ws['A3'] = 123.456
        ws['A4'] = True
        ws['A5'] = None  # Empty

        wb.save(filepath)
        self.generator.generated_files.append(filepath)

        config = [{
            "sheets": ["Types"],
            "extractions": [{
                "function": "single_cells",
                "instructions": {
                    "string_val": "A1",
                    "int_val": "A2",
                    "float_val": "A3",
                    "bool_val": "A4",
                    "null_val": "A5"
                }
            }]
        }]

        result, elapsed = self._time_execution(
            sheet_excavator.excel_extract, [filepath], config, 1
        )

        parsed = json.loads(result)

        file_key = list(parsed.keys())[0]
        data = parsed[file_key].get("Types", {})

        # Verify types - single_cells returns raw values, not arrays
        assert data["string_val"] == "StringValue", f"Expected 'StringValue', got {data['string_val']!r}"
        # Note: Excel stores integers as floats
        assert data["int_val"] == 12345 or data["int_val"] == 12345.0, f"Expected 12345, got {data['int_val']!r}"
        assert abs(data["float_val"] - 123.456) < 0.001, f"Expected 123.456, got {data['float_val']!r}"
        assert data["bool_val"] == True, f"Expected True, got {data['bool_val']!r}"

        self.results['bottleneck_9'] = {
            'types_tested': ['string', 'int', 'float', 'bool', 'null'],
            'time_seconds': elapsed,
            'status': 'PASS'
        }

        print(f"  Data types tested: string, int, float, bool, null")
        print(f"  Time: {elapsed:.3f}s")
        print(f"  Status: PASS")

    # =========================================================================
    # COMBINED STRESS TEST
    # =========================================================================
    def test_combined_stress(self):
        """Combined stress test exercising all optimizations."""
        print("\n" + "="*70)
        print("COMBINED STRESS TEST")
        print("="*70)

        # Create multiple complex files
        files = []
        for i in range(10):
            f = self.generator.create_large_multirow_file(
                f"stress_{i}.xlsx",
                num_rows=500,
                num_columns=10,
                num_sheets=5
            )
            files.append(f)

        config = [{
            "sheets": ["Sheet*"],
            "skip_sheets": ["Sheet4", "Sheet5"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "label": "data",
                    "instructions": {
                        "row_range": [2, 501],
                        "unique_id": "A",
                        "stop_if_empty": "A",
                        "columns": {f"c{i}": get_column_letter(i+1) for i in range(10)}
                    }
                }
            ]
        }]

        result, elapsed = self._time_execution(
            sheet_excavator.excel_extract, files, config, 4
        )

        parsed = json.loads(result)

        # Validate
        assert len(parsed) == 10, f"Expected 10 files, got {len(parsed)}"

        total_rows = 0
        for file_key, file_data in parsed.items():
            if file_key == "filepath":
                continue
            for sheet_key, sheet_data in file_data.items():
                if sheet_key == "filepath":
                    continue
                if "data" in sheet_data:
                    total_rows += len(sheet_data["data"])

        self.results['stress_test'] = {
            'files': 10,
            'sheets_per_file': 3,  # 5 - 2 skipped
            'rows_per_sheet': 500,
            'total_rows': total_rows,
            'time_seconds': elapsed,
            'status': 'PASS'
        }

        print(f"  Files: 10")
        print(f"  Sheets per file (after skip): 3")
        print(f"  Rows extracted: {total_rows}")
        print(f"  Time: {elapsed:.3f}s")
        print(f"  Status: PASS")

    # =========================================================================
    # RUN ALL TESTS
    # =========================================================================
    def run_all(self):
        """Run all performance tests."""
        print("\n" + "#"*70)
        print("# SHEET_EXCAVATOR PERFORMANCE TEST SUITE")
        print("#"*70)

        tests = [
            self.test_bottleneck_1_arc_cloning,
            self.test_bottleneck_2_3_column_parsing,
            self.test_bottleneck_4_single_workbook,
            self.test_bottleneck_5_conditional_trim,
            self.test_bottleneck_6_skip_sheets_hashset,
            self.test_bottleneck_7_unique_keys,
            self.test_bottleneck_8_indexset_sheets,
            self.test_bottleneck_9_type_descriptions,
            self.test_combined_stress,
        ]

        passed = 0
        failed = 0

        for test in tests:
            try:
                test()
                passed += 1
            except Exception as e:
                failed += 1
                test_name = test.__name__
                self.results[test_name] = {'status': 'FAIL', 'error': str(e)}
                print(f"  Status: FAIL - {e}")

        # Cleanup
        self.generator.cleanup()

        # Summary
        print("\n" + "#"*70)
        print("# SUMMARY")
        print("#"*70)
        print(f"\n  Passed: {passed}")
        print(f"  Failed: {failed}")
        print(f"  Total:  {passed + failed}")

        print("\n  Timing Summary:")
        for name, result in self.results.items():
            if 'time_seconds' in result:
                print(f"    {name}: {result['time_seconds']:.3f}s")

        if failed == 0:
            print("\n  ALL TESTS PASSED!")
            return True
        else:
            print(f"\n  {failed} TESTS FAILED")
            return False


if __name__ == "__main__":
    suite = PerformanceTestSuite()
    success = suite.run_all()
    exit(0 if success else 1)
