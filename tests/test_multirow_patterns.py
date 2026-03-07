"""Tests for multirow_patterns extraction function."""

from conftest import run_extract


def test_single_unique_id(tmp_path):
    """Dict result keyed by unique_id column."""
    from openpyxl import Workbook

    filepath = tmp_path / "single_uid.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    data = [
        ("ID001", "Project Alpha", 50000, 0.8),
        ("ID002", "Project Beta", 30000, 0.6),
        ("ID003", "Project Gamma", 10000, 0.4),
    ]
    for row_idx, (title, desc, est, chance) in enumerate(data, start=1):
        ws.cell(row=row_idx, column=2, value=title)
        ws.cell(row=row_idx, column=3, value=desc)
        ws.cell(row=row_idx, column=4, value=est)
        ws.cell(row=row_idx, column=5, value=chance)
    wb.save(filepath)

    config = [
        {
            "sheets": ["Sheet 1"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "label": "deposits",
                    "instructions": {
                        "row_range": [1, 10],
                        "unique_id": "B",
                        "columns": {"Title": "B", "Description": "C", "Estimate": "D", "Chance": "E"},
                    },
                }
            ],
        }
    ]
    result = run_extract([str(filepath)], config)
    file_key = list(result.keys())[0]
    deposits = result[file_key]["Sheet 1"]["deposits"]

    assert isinstance(deposits, dict)
    assert len(deposits) == 3
    assert "ID001" in deposits
    assert deposits["ID001"]["Estimate"] == 50000.0


def test_composite_unique_id(tmp_excel_composite):
    """Composite key from ["B","C"]."""
    config = [
        {
            "sheets": ["Sheet 1"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "label": "projects",
                    "instructions": {
                        "row_range": [1, 50],
                        "unique_id": ["B", "C"],
                        "columns": {"Project": "B", "Year": "C", "Budget": "D", "Status": "E"},
                    },
                }
            ],
        }
    ]
    result = run_extract([tmp_excel_composite], config)
    file_key = list(result.keys())[0]
    projects = result[file_key]["Sheet 1"]["projects"]

    assert isinstance(projects, dict)
    assert len(projects) == 4


def test_composite_custom_separator(tmp_excel_composite):
    """Custom separator '-'."""
    config = [
        {
            "sheets": ["Sheet 1"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "label": "projects",
                    "instructions": {
                        "row_range": [1, 50],
                        "unique_id": ["B", "C"],
                        "unique_id_separator": "-",
                        "columns": {"Project": "B", "Year": "C", "Budget": "D", "Status": "E"},
                    },
                }
            ],
        }
    ]
    result = run_extract([tmp_excel_composite], config)
    file_key = list(result.keys())[0]
    projects = result[file_key]["Sheet 1"]["projects"]

    assert isinstance(projects, dict)
    # Keys should contain '-' separator
    keys = list(projects.keys())
    assert any("-" in k for k in keys)


def test_no_unique_id_returns_array(tmp_excel_with_gaps):
    """Array result when unique_id omitted."""
    config = [
        {
            "sheets": ["TestSheet"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "label": "items",
                    "instructions": {
                        "row_range": [1, 1000],
                        "stop_if_empty": "A",
                        "columns": {"ID": "A", "Name": "B", "Value": "C"},
                    },
                }
            ],
        }
    ]
    result = run_extract([tmp_excel_with_gaps], config)
    file_key = list(result.keys())[0]
    items = result[file_key]["TestSheet"]["items"]

    assert isinstance(items, list)
    assert len(items) == 3
    assert items[0]["ID"] == "001"


def test_stop_if_empty_column(tmp_excel_with_gaps):
    """Stop when single column empty."""
    config = [
        {
            "sheets": ["TestSheet"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "label": "items",
                    "instructions": {
                        "row_range": [1, 100],
                        "stop_if_empty": "A",
                        "columns": {"ID": "A", "Name": "B", "Value": "C"},
                    },
                }
            ],
        }
    ]
    result = run_extract([tmp_excel_with_gaps], config)
    file_key = list(result.keys())[0]
    items = result[file_key]["TestSheet"]["items"]

    assert isinstance(items, list)
    assert len(items) == 3


def test_stop_if_empty_row_mode(tmp_path):
    """Stop when entire row empty."""
    from openpyxl import Workbook

    filepath = tmp_path / "row_mode.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    ws["A1"] = "ID001"
    ws["B1"] = "Name1"
    ws["C1"] = 100
    # Row 2: A empty but B has data
    ws["B2"] = "Name2"
    ws["C2"] = 200
    ws["A3"] = "ID003"
    ws["B3"] = "Name3"
    ws["C3"] = 300
    # Row 4 completely empty - should stop
    wb.save(filepath)

    config = [
        {
            "sheets": ["Sheet 1"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "label": "records",
                    "instructions": {
                        "row_range": [1, 500],
                        "stop_if_empty": {"mode": "row", "consecutive": 1},
                        "columns": {"Field1": "A", "Field2": "B", "Field3": "C"},
                    },
                }
            ],
        }
    ]
    result = run_extract([str(filepath)], config)
    file_key = list(result.keys())[0]
    records = result[file_key]["Sheet 1"]["records"]

    assert isinstance(records, list)
    assert len(records) == 3


def test_stop_if_empty_multiple_columns(tmp_path):
    """Stop when ALL specified columns empty."""
    from datetime import datetime

    from openpyxl import Workbook

    filepath = tmp_path / "multi_col_stop.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    ws["A1"] = "ID001"
    ws["B1"] = datetime(2024, 1, 15)
    ws["C1"] = 100
    # Row 2: A empty but B has data
    ws["B2"] = datetime(2024, 2, 15)
    ws["C2"] = 200
    ws["A3"] = "ID003"
    # B3 empty but A has data
    ws["C3"] = 300
    # Row 4: both A and B empty - should stop
    wb.save(filepath)

    config = [
        {
            "sheets": ["Sheet 1"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "label": "transactions",
                    "instructions": {
                        "row_range": [1, 1000],
                        "unique_id": "A",
                        "stop_if_empty": ["A", "B"],
                        "columns": {"ID": "A", "Date": "B", "Amount": "C"},
                    },
                }
            ],
        }
    ]
    result = run_extract([str(filepath)], config)
    file_key = list(result.keys())[0]
    transactions = result[file_key]["Sheet 1"]["transactions"]

    assert isinstance(transactions, dict)


def test_stop_if_empty_object_syntax(tmp_excel_with_gaps):
    """Object syntax: {"column": "A", "consecutive": 1}."""
    config = [
        {
            "sheets": ["TestSheet"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "label": "items",
                    "instructions": {
                        "row_range": [1, 100],
                        "stop_if_empty": {"column": "A", "consecutive": 1},
                        "columns": {"ID": "A", "Name": "B", "Value": "C"},
                    },
                }
            ],
        }
    ]
    result = run_extract([tmp_excel_with_gaps], config)
    file_key = list(result.keys())[0]
    items = result[file_key]["TestSheet"]["items"]

    assert isinstance(items, list)
    assert len(items) == 3


def test_stop_consecutive_gap_tolerance(tmp_excel_with_gaps):
    """Gap tolerance with stop_consecutive: 2 (tolerate 1 gap)."""
    config = [
        {
            "sheets": ["TestSheet"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "label": "with_gaps",
                    "instructions": {
                        "row_range": [1, 100],
                        "stop_if_empty": "A",
                        "stop_consecutive": 2,
                        "columns": {"ID": "A", "Name": "B", "Value": "C"},
                    },
                }
            ],
        }
    ]
    result = run_extract([tmp_excel_with_gaps], config)
    file_key = list(result.keys())[0]
    items = result[file_key]["TestSheet"]["with_gaps"]

    assert isinstance(items, list)
    # Should get 4 items (001, 002, 003, 005) - tolerates 1 gap
    assert len(items) == 4


def test_duplicate_key_dedup(tmp_excel_duplicates):
    """_1, _2 suffix dedup for duplicate keys."""
    config = [
        {
            "sheets": ["Duplicates"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "instructions": {
                        "row_range": [1, 100],
                        "unique_id": "A",
                        "columns": {"value": "B", "data": "C"},
                    },
                }
            ],
        }
    ]
    result = run_extract([tmp_excel_duplicates], config)
    file_key = list(result.keys())[0]
    data = result[file_key]["Duplicates"]

    assert len(data) == 100
    assert "DUPLICATE_KEY" in data
    assert "DUPLICATE_KEY_1" in data
    assert "DUPLICATE_KEY_99" in data


def test_column_merge(tmp_path):
    """Multiple columns per field ["X","Y","Z"]."""
    from openpyxl import Workbook

    filepath = tmp_path / "col_merge.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "ID001"
    ws["B1"] = "ValB"
    ws["C1"] = "ValC"
    ws["D1"] = "ValD"
    ws["A2"] = "ID002"
    ws["B2"] = "ValB2"
    # C2 and D2 empty
    wb.save(filepath)

    config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "label": "data",
                    "instructions": {
                        "row_range": [1, 10],
                        "unique_id": "A",
                        "stop_if_empty": "A",
                        "columns": {"ID": "A", "Combined": ["B", "C", "D"]},
                    },
                }
            ],
        }
    ]
    result = run_extract([str(filepath)], config)
    file_key = list(result.keys())[0]
    data = result[file_key]["Sheet1"]["data"]

    assert isinstance(data, dict)
    # First row should have array of 3 values
    assert isinstance(data["ID001"]["Combined"], list)
    assert len(data["ID001"]["Combined"]) == 3
