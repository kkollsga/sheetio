"""Tests for ExtractionConfig builder class."""

import json

import pytest
from conftest import run_extract
from openpyxl import Workbook
from sheetio import ExtractionConfig

# ---------------------------------------------------------------------------
# Unit tests: verify .build() output (no Excel files needed)
# ---------------------------------------------------------------------------


def test_empty_config():
    """Empty config builds to empty list."""
    assert ExtractionConfig().build() == []


def test_single_cells_kwargs():
    """kwargs produce correct single_cells instructions."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet1"]).single_cells("header", title="B2", date="D4")
    built = config.build()

    assert len(built) == 1
    ext = built[0]["extractions"][0]
    assert ext["function"] == "single_cells"
    assert ext["label"] == "header"
    assert ext["instructions"] == {"title": "B2", "date": "D4"}


def test_single_cells_cells_dict():
    """cells= dict produces correct instructions."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet1"]).single_cells("meta", cells={"Report Title": "B2", "Date (ISO)": "D4"})
    built = config.build()

    ext = built[0]["extractions"][0]
    assert ext["instructions"]["Report Title"] == "B2"
    assert ext["instructions"]["Date (ISO)"] == "D4"


def test_single_cells_mixed():
    """Both cells= and kwargs are merged, kwargs take precedence."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet1"]).single_cells("header", cells={"title": "A1", "other": "C3"}, title="B2")
    built = config.build()

    instr = built[0]["extractions"][0]["instructions"]
    assert instr["title"] == "B2"  # kwargs wins
    assert instr["other"] == "C3"


def test_multirow_build():
    """multirow with all options builds correctly."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet1"]).multirow(
        "items",
        row_range=(10, 100),
        unique_id=["A", "B"],
        unique_id_separator="-",
        stop_if_empty="A",
        stop_consecutive=3,
        columns={"ID": "A", "Name": "B", "Combined": ["X", "Y"]},
    )
    built = config.build()

    ext = built[0]["extractions"][0]
    assert ext["function"] == "multirow_patterns"
    assert ext["label"] == "items"
    instr = ext["instructions"]
    assert instr["row_range"] == [10, 100]
    assert instr["unique_id"] == ["A", "B"]
    assert instr["unique_id_separator"] == "-"
    assert instr["stop_if_empty"] == "A"
    assert instr["stop_consecutive"] == 3
    assert instr["columns"]["Combined"] == ["X", "Y"]


def test_multirow_minimal():
    """multirow with only required fields omits optional keys."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet1"]).multirow(row_range=(1, 50), columns={"val": "A"})
    built = config.build()

    instr = built[0]["extractions"][0]["instructions"]
    assert "unique_id" not in instr
    assert "stop_if_empty" not in instr
    assert "stop_consecutive" not in instr
    assert "unique_id_separator" not in instr


def test_dataframe_build():
    """dataframe with all options builds correctly."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet1"]).dataframe(
        "table",
        row_range=(5, 50),
        column_range=["B", "F"],
        header_row=[2, 3, 4],
        separator="|",
    )
    built = config.build()

    ext = built[0]["extractions"][0]
    assert ext["function"] == "dataframe"
    assert ext["label"] == "table"
    instr = ext["instructions"]
    assert instr["row_range"] == [5, 50]
    assert instr["column_range"] == ["B", "F"]
    assert instr["header_row"] == [2, 3, 4]
    assert instr["separator"] == "|"


def test_duplicate_label_overwrites():
    """Second extraction with same label replaces the first."""
    config = ExtractionConfig()
    group = config.add_sheets(["Sheet1"])
    group.single_cells("header", title="A1")
    group.single_cells("header", title="B2")
    built = config.build()

    assert len(built[0]["extractions"]) == 1
    assert built[0]["extractions"][0]["instructions"]["title"] == "B2"


def test_method_chaining_returns_self():
    """Each extraction method returns the SheetGroup for chaining."""
    config = ExtractionConfig()
    group = config.add_sheets(["Sheet1"])
    result = group.single_cells("a", title="A1")
    assert result is group
    result = group.multirow("b", row_range=(1, 10), columns={"v": "A"})
    assert result is group
    result = group.dataframe("c", row_range=(1, 10), column_range=["A", "B"], header_row=1)
    assert result is group


def test_cross_group_chaining():
    """add_sheets() on SheetGroup creates new group on parent."""
    config = ExtractionConfig()
    group1 = config.add_sheets(["Sheet1"])
    group2 = group1.add_sheets(["Sheet2"])
    assert group2 is not group1
    assert len(config._groups) == 2


def test_remove_extraction():
    """remove() deletes extraction by label."""
    config = ExtractionConfig()
    group = config.add_sheets(["Sheet1"])
    group.single_cells("header", title="A1")
    group.single_cells("footer", title="A100")
    group.remove("header")
    built = config.build()

    assert len(built[0]["extractions"]) == 1
    assert built[0]["extractions"][0]["label"] == "footer"


def test_remove_nonexistent_raises():
    """remove() raises KeyError for unknown label."""
    config = ExtractionConfig()
    group = config.add_sheets(["Sheet1"])
    with pytest.raises(KeyError, match="missing"):
        group.remove("missing")


def test_optional_fields_omitted():
    """None-valued optional fields are not in build output."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet1"]).single_cells(title="A1")
    built = config.build()

    ext = built[0]["extractions"][0]
    assert "label" not in ext
    assert "break_if_null" not in ext
    assert "skip_sheets" not in built[0]


def test_skip_sheets_in_build():
    """skip_sheets appears only when set."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet*"], skip=["Sheet3"]).single_cells(title="A1")
    built = config.build()

    assert built[0]["skip_sheets"] == ["Sheet3"]


def test_break_if_null_at_group_level():
    """break_if_null at group level appears in group dict."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet1"], break_if_null="C3").single_cells(title="A1")
    built = config.build()

    assert built[0]["break_if_null"] == "C3"


def test_break_if_null_at_extraction_level():
    """break_if_null at extraction level appears in extraction dict."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet1"]).single_cells("header", title="A1", break_if_null="C3")
    built = config.build()

    assert built[0]["extractions"][0]["break_if_null"] == "C3"


def test_row_range_tuple_accepted():
    """tuple (1, 50) works same as list [1, 50]."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet1"]).multirow(row_range=(1, 50), columns={"v": "A"})
    built = config.build()

    assert built[0]["extractions"][0]["instructions"]["row_range"] == [1, 50]


# ---------------------------------------------------------------------------
# Integration tests: verify actual extraction with Excel files
# ---------------------------------------------------------------------------


def test_extract_returns_dict(tmp_excel_simple):
    """extract() returns a parsed dict, not a JSON string."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet1"]).single_cells("data", title="A1")
    result = config.extract([tmp_excel_simple])

    assert isinstance(result, dict)
    file_key = list(result.keys())[0]
    assert result[file_key]["Sheet1"]["data"]["title"] == "Title Value"


def test_extract_matches_raw_api(tmp_excel_simple):
    """Builder result matches manual config passed to excel_extract."""
    manual_config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [
                {"function": "single_cells", "label": "data", "instructions": {"title": "A1", "number": "C3"}}
            ],
        }
    ]
    manual_result = run_extract([tmp_excel_simple], manual_config)

    config = ExtractionConfig()
    config.add_sheets(["Sheet1"]).single_cells("data", title="A1", number="C3")
    builder_result = config.extract([tmp_excel_simple])

    assert builder_result == manual_result


def test_full_workflow(tmp_path):
    """Build config, extract, verify data correctness."""
    filepath = tmp_path / "workflow.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Report"
    ws["B1"] = 2024
    for row in range(3, 8):
        ws.cell(row=row, column=1, value=f"ID_{row - 2}")
        ws.cell(row=row, column=2, value=(row - 2) * 100)
    wb.save(filepath)

    config = ExtractionConfig()
    group = config.add_sheets(["Data"])
    group.single_cells("header", title="A1", year="B1")
    group.multirow(
        "items",
        row_range=(3, 7),
        unique_id="A",
        columns={"id": "A", "value": "B"},
    )

    result = config.extract([str(filepath)])
    file_key = list(result.keys())[0]
    data = result[file_key]["Data"]

    assert data["header"]["title"] == "Report"
    assert data["header"]["year"] == 2024.0 or data["header"]["year"] == 2024
    assert len(data["items"]) == 5
    assert "ID_1" in data["items"]


# ---------------------------------------------------------------------------
# Serialization tests
# ---------------------------------------------------------------------------


def test_to_json_from_json_roundtrip(tmp_path):
    """Save to JSON, load back, build() produces identical output."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet1"], skip=["Summary"]).single_cells("header", title="B2").multirow(
        "items", row_range=(10, 100), unique_id="A", columns={"ID": "A", "Name": "B"}
    )
    config.add_sheets(["Data_*"]).dataframe("table", header_row=[2, 3], row_range=(5, 50), column_range=["B", "F"])

    json_path = str(tmp_path / "config.json")
    config.to_json(json_path)
    loaded = ExtractionConfig.from_json(json_path)

    assert config.build() == loaded.build()


def test_from_json_manual_config(tmp_path, tmp_excel_simple):
    """Load hand-written JSON config, verify extraction works."""
    manual = [
        {
            "sheets": ["Sheet1"],
            "extractions": [{"function": "single_cells", "label": "data", "instructions": {"title": "A1"}}],
        }
    ]
    json_path = str(tmp_path / "manual.json")
    with open(json_path, "w") as f:
        json.dump(manual, f)

    config = ExtractionConfig.from_json(json_path)
    result = config.extract([tmp_excel_simple])

    file_key = list(result.keys())[0]
    assert result[file_key]["Sheet1"]["data"]["title"] == "Title Value"


# ---------------------------------------------------------------------------
# Summary test
# ---------------------------------------------------------------------------


def test_summary_output(capsys):
    """summary() prints human-readable overview."""
    config = ExtractionConfig()
    config.add_sheets(["Sheet1"]).single_cells("header", title="B2")
    config.add_sheets(["Data_*"], skip=["Data_X"]).multirow(
        "items", row_range=(1, 50), unique_id="A", columns={"ID": "A"}
    )
    config.summary()

    captured = capsys.readouterr().out
    assert "2 sheet groups" in captured
    assert "Sheet1" in captured
    assert "single_cells" in captured
    assert "Data_*" in captured
    assert "multirow_patterns" in captured


# ---------------------------------------------------------------------------
# Backward compatibility
# ---------------------------------------------------------------------------


def test_existing_import_still_works():
    """from sheetio import excel_extract still works."""
    from sheetio import excel_extract

    assert callable(excel_extract)
    result = json.loads(excel_extract([], [{"sheets": ["S"], "extractions": []}], 1))
    assert isinstance(result, dict)
