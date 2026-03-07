"""Tests for edge cases and advanced features."""

import pytest
from conftest import run_extract


def test_wildcard_sheet_matching(tmp_excel_multisheet):
    """'School_*' matches School_A, School_B, School_C only."""
    config = [
        {
            "sheets": ["School_*"],
            "extractions": [{"function": "single_cells", "instructions": {"data": "A1"}}],
        }
    ]
    result = run_extract([tmp_excel_multisheet], config)
    file_key = list(result.keys())[0]
    sheets = [k for k in result[file_key].keys() if k != "filepath"]

    assert len(sheets) == 3
    assert "Other_Sheet" not in sheets


def test_skip_sheets(tmp_path):
    """skip_sheets excludes specified sheets."""
    from openpyxl import Workbook

    filepath = tmp_path / "skip.xlsx"
    wb = Workbook()
    for i, name in enumerate(["Data_01", "Data_02", "Data_03", "Data_Summary"]):
        if i == 0:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(name)
        ws["A1"] = f"Content of {name}"
    wb.save(filepath)

    config = [
        {
            "sheets": ["Data_*"],
            "skip_sheets": ["Data_Summary"],
            "extractions": [{"function": "single_cells", "instructions": {"content": "A1"}}],
        }
    ]
    result = run_extract([str(filepath)], config)
    file_key = list(result.keys())[0]
    sheets = [k for k in result[file_key].keys() if k != "filepath"]

    assert len(sheets) == 3
    assert "Data_Summary" not in sheets


def test_multiple_extractions_same_sheet(tmp_path):
    """single_cells + multirow combined on same sheet."""
    from datetime import datetime

    from openpyxl import Workbook

    filepath = tmp_path / "multi_extract.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws["A1"] = "Report Title"
    ws["B1"] = "Q4 2024 Summary"
    ws["A2"] = "Generated"
    ws["B2"] = datetime(2024, 12, 31)
    for i in range(5, 10):
        ws.cell(row=i, column=1, value=f"ITEM{i - 4}")
        ws.cell(row=i, column=2, value=(i - 4) * 100)
    wb.save(filepath)

    config = [
        {
            "sheets": ["Report"],
            "extractions": [
                {"function": "single_cells", "label": "header", "instructions": {"title": "B1", "date": "B2"}},
                {
                    "function": "multirow_patterns",
                    "label": "data",
                    "instructions": {
                        "row_range": [5, 20],
                        "unique_id": "A",
                        "stop_if_empty": "A",
                        "columns": {"ID": "A", "Value": "B"},
                    },
                },
            ],
        }
    ]
    result = run_extract([str(filepath)], config)
    file_key = list(result.keys())[0]
    report = result[file_key]["Report"]

    assert "header" in report
    assert "data" in report
    assert report["header"]["title"] == "Q4 2024 Summary"
    assert len(report["data"]) == 5


def test_break_if_null_allows_extraction(tmp_path):
    """Extraction proceeds when break_if_null cell has value."""
    from openpyxl import Workbook

    filepath = tmp_path / "break_ok.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Value A"
    ws["C3"] = "Not Null"
    wb.save(filepath)

    config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [
                {
                    "function": "single_cells",
                    "break_if_null": "c3",
                    "instructions": {"test": "a1"},
                }
            ],
        }
    ]
    result = run_extract([str(filepath)], config)
    file_key = list(result.keys())[0]
    assert result[file_key]["Sheet1"]["test"] == "Value A"


def test_empty_sheet(tmp_path):
    """Sheet exists but has no data."""
    from openpyxl import Workbook

    filepath = tmp_path / "empty.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Empty"
    wb.save(filepath)

    config = [
        {
            "sheets": ["Empty"],
            "extractions": [
                {
                    "function": "single_cells",
                    "instructions": {"val": "a1"},
                }
            ],
        }
    ]
    result = run_extract([str(filepath)], config)
    assert isinstance(result, dict)


@pytest.mark.slow
def test_parallel_processing(tmp_path):
    """Multiple files with workers=4."""
    from datetime import datetime

    from openpyxl import Workbook

    files = []
    for i in range(10):
        filepath = tmp_path / f"parallel_{i:03d}.xlsx"
        wb = Workbook()
        for sheet_idx in range(3):
            if sheet_idx == 0:
                ws = wb.active
                ws.title = f"Data_{sheet_idx + 1}"
            else:
                ws = wb.create_sheet(f"Data_{sheet_idx + 1}")
            ws["A1"] = f"Report {i + 1}"
            ws["B1"] = datetime.now()
            for row in range(5, 25):
                ws.cell(row=row, column=1, value=f"ID_{i}_{sheet_idx}_{row}")
                ws.cell(row=row, column=2, value=f"Name_{row}")
                ws.cell(row=row, column=3, value=row * 10.5)
        wb.save(filepath)
        files.append(str(filepath))

    config = [
        {
            "sheets": ["Data_*"],
            "extractions": [
                {"function": "single_cells", "label": "header", "instructions": {"title": "A1"}},
                {
                    "function": "multirow_patterns",
                    "label": "records",
                    "instructions": {
                        "row_range": [5, 100],
                        "unique_id": "A",
                        "stop_if_empty": "A",
                        "columns": {"ID": "A", "Name": "B", "Value": "C"},
                    },
                },
            ],
        }
    ]
    result = run_extract(files, config, workers=4)
    assert len(result) == 10
