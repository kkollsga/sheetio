"""Performance and scale tests ported from the original performance_test.py.

These tests verify correctness at scale -- they exercise the same code paths
as the unit tests but with larger datasets that stress optimizations like
Arc sharing, pre-parsed configs, HashSet skip_sheets, and IndexSet sheet matching.
"""

import json

import pytest
import sheetio
from conftest import run_extract
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _create_large_multirow_file(tmp_path, filename, num_rows=1000, num_columns=15, num_sheets=3):
    """Create Excel file with many rows/columns and mixed data types."""
    filepath = tmp_path / filename
    wb = Workbook()
    for sheet_idx in range(num_sheets):
        if sheet_idx == 0:
            ws = wb.active
            ws.title = f"Sheet{sheet_idx + 1}"
        else:
            ws = wb.create_sheet(f"Sheet{sheet_idx + 1}")
        for col in range(1, num_columns + 1):
            ws.cell(row=1, column=col, value=f"Column_{get_column_letter(col)}")
        for row in range(2, num_rows + 2):
            for col in range(1, num_columns + 1):
                if col == 1:
                    ws.cell(row=row, column=col, value=f"ID_{row - 1}")
                elif col % 3 == 0:
                    ws.cell(row=row, column=col, value=row * col * 1.5)
                elif col % 3 == 1:
                    ws.cell(row=row, column=col, value=row * col)
                else:
                    ws.cell(row=row, column=col, value=f"  Data_{row}_{col}  ")
    wb.save(filepath)
    return str(filepath)


def _create_many_sheets_file(tmp_path, filename, num_sheets=50):
    """Create file with many sheets."""
    filepath = tmp_path / filename
    wb = Workbook()
    for i in range(num_sheets):
        if i == 0:
            ws = wb.active
            ws.title = f"Sheet_{i + 1:03d}"
        else:
            ws = wb.create_sheet(f"Sheet_{i + 1:03d}")
        ws["A1"] = f"Sheet {i + 1}"
        ws["B1"] = i * 100
    wb.save(filepath)
    return str(filepath)


# ---------------------------------------------------------------------------
# Data type handling
# ---------------------------------------------------------------------------
def test_all_data_types(tmp_path):
    """Verify string, int, float, bool, and null are extracted with correct types."""
    filepath = tmp_path / "datatypes.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Types"
    ws["A1"] = "StringValue"
    ws["A2"] = 12345
    ws["A3"] = 123.456
    ws["A4"] = True
    # A5 left empty (null)
    wb.save(filepath)

    config = [
        {
            "sheets": ["Types"],
            "extractions": [
                {
                    "function": "single_cells",
                    "instructions": {
                        "string_val": "A1",
                        "int_val": "A2",
                        "float_val": "A3",
                        "bool_val": "A4",
                        "null_val": "A5",
                    },
                }
            ],
        }
    ]
    result = run_extract([str(filepath)], config)
    file_key = list(result.keys())[0]
    data = result[file_key]["Types"]

    assert data["string_val"] == "StringValue"
    assert data["int_val"] == 12345 or data["int_val"] == 12345.0
    assert abs(data["float_val"] - 123.456) < 0.001
    assert data["bool_val"] is True
    assert data["null_val"] is None


# ---------------------------------------------------------------------------
# Whitespace trimming
# ---------------------------------------------------------------------------
def test_whitespace_trimming(tmp_path):
    """Cells with leading/trailing whitespace are trimmed in multirow unique_id."""
    filepath = tmp_path / "trim.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in range(1, 51):
        ws.cell(row=row, column=1, value=f"  ID_{row}  ")  # Whitespace padding
        ws.cell(row=row, column=2, value=f"  Data_{row}  ")
    wb.save(filepath)

    config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "instructions": {
                        "row_range": [1, 50],
                        "unique_id": "A",
                        "columns": {"id": "A", "data": "B"},
                    },
                }
            ],
        }
    ]
    result = run_extract([str(filepath)], config)
    file_key = list(result.keys())[0]
    data = result[file_key]["Sheet1"]

    # Unique IDs should be trimmed (used as dict keys)
    assert "ID_1" in data
    assert "ID_50" in data
    assert len(data) == 50


# ---------------------------------------------------------------------------
# Large-scale multirow extraction (2000 rows x 15 columns)
# ---------------------------------------------------------------------------
@pytest.mark.slow
def test_large_multirow_extraction(tmp_path):
    """Extract 2000 rows x 15 columns to verify pre-parsed config optimization."""
    filepath = _create_large_multirow_file(tmp_path, "large.xlsx", num_rows=2000, num_columns=15, num_sheets=1)

    columns = {f"col_{get_column_letter(i)}": get_column_letter(i) for i in range(1, 16)}
    config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "instructions": {"row_range": [2, 2001], "unique_id": "A", "columns": columns},
                }
            ],
        }
    ]
    result = run_extract([filepath], config)
    file_key = list(result.keys())[0]
    sheet_data = result[file_key]["Sheet1"]

    assert len(sheet_data) == 2000
    first_key = list(sheet_data.keys())[0]
    assert len(sheet_data[first_key]) == 15


# ---------------------------------------------------------------------------
# All 3 extraction types on same file
# ---------------------------------------------------------------------------
def test_three_extraction_types_on_same_file(tmp_path):
    """single_cells + multirow_patterns + dataframe on a single sheet."""
    filepath = tmp_path / "multi_type.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Section 1: Single cells
    ws["A1"] = "Header1"
    ws["B1"] = "Value1"
    ws["A2"] = "Header2"
    ws["B2"] = 12345

    # Section 2: Multirow data (rows 5-54)
    for row in range(5, 55):
        ws.cell(row=row, column=1, value=f"Row{row - 4}")
        ws.cell(row=row, column=2, value=row * 10)
        ws.cell(row=row, column=3, value=f"Text_{row}")

    # Section 3: Dataframe (header row 60, data 61-79, columns E-I)
    for col in range(5, 10):
        ws.cell(row=60, column=col, value=f"Col{col - 4}")
        for row in range(61, 80):
            ws.cell(row=row, column=col, value=row + col)

    wb.save(filepath)

    config = [
        {
            "sheets": ["Data"],
            "extractions": [
                {
                    "function": "single_cells",
                    "label": "header_data",
                    "instructions": {"cell1": "A1", "cell2": "B1", "cell3": "A2", "cell4": "B2"},
                },
                {
                    "function": "multirow_patterns",
                    "label": "row_data",
                    "instructions": {
                        "row_range": [5, 54],
                        "unique_id": "A",
                        "columns": {"name": "A", "value": "B", "text": "C"},
                    },
                },
                {
                    "function": "dataframe",
                    "label": "table_data",
                    "instructions": {"header_row": 60, "row_range": [61, 79], "column_range": ["E", "I"]},
                },
            ],
        }
    ]
    result = run_extract([str(filepath)], config)
    file_key = list(result.keys())[0]
    sheet_data = result[file_key]["Data"]

    assert "header_data" in sheet_data
    assert "row_data" in sheet_data
    assert "table_data" in sheet_data

    assert sheet_data["header_data"]["cell1"] == "Header1"
    assert sheet_data["header_data"]["cell4"] == 12345 or sheet_data["header_data"]["cell4"] == 12345.0
    assert len(sheet_data["row_data"]) == 50
    assert isinstance(sheet_data["table_data"], dict)


# ---------------------------------------------------------------------------
# Many sheets with large skip_sheets (50 sheets, skip 48)
# ---------------------------------------------------------------------------
def test_skip_sheets_at_scale(tmp_path):
    """50 sheets, skip 48 -- verifies HashSet optimization for skip_sheets."""
    filepath = _create_many_sheets_file(tmp_path, "many_sheets.xlsx", num_sheets=50)

    skip_list = [f"Sheet_{i:03d}" for i in range(3, 51)]
    config = [
        {
            "sheets": ["Sheet_*"],
            "skip_sheets": skip_list,
            "extractions": [{"function": "single_cells", "instructions": {"value": "A1", "number": "B1"}}],
        }
    ]
    result = run_extract([filepath], config)
    file_key = list(result.keys())[0]
    sheets_processed = [k for k in result[file_key].keys() if k != "filepath"]

    assert len(sheets_processed) == 2


# ---------------------------------------------------------------------------
# Wildcard matching 50 sheets
# ---------------------------------------------------------------------------
def test_wildcard_matching_many_sheets(tmp_path):
    """Wildcard matches all 50 sheets -- verifies IndexSet optimization."""
    filepath = _create_many_sheets_file(tmp_path, "indexset.xlsx", num_sheets=50)

    config = [
        {
            "sheets": ["Sheet_*"],
            "extractions": [{"function": "single_cells", "instructions": {"title": "A1"}}],
        }
    ]
    result = run_extract([filepath], config)
    file_key = list(result.keys())[0]
    sheets_processed = [k for k in result[file_key].keys() if k != "filepath"]

    assert len(sheets_processed) == 50


# ---------------------------------------------------------------------------
# Bulk file processing (50 files, 8 workers)
# ---------------------------------------------------------------------------
@pytest.mark.slow
def test_bulk_file_processing(tmp_path):
    """Process 50 files with 8 workers -- verifies Arc sharing optimization."""
    files = []
    for i in range(50):
        filepath = tmp_path / f"bulk_{i + 1:03d}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        for row in range(1, 21):
            ws.cell(row=row, column=1, value=f"ID_{row}")
            ws.cell(row=row, column=2, value=row * 10)
            ws.cell(row=row, column=3, value=f"Value_{row}")
        wb.save(filepath)
        files.append(str(filepath))

    config = [
        {
            "sheets": ["Data"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "instructions": {
                        "row_range": [1, 20],
                        "unique_id": "A",
                        "columns": {f"col_{i}": chr(65 + i) for i in range(3)},
                    },
                }
            ]
            * 5,  # 5 identical extractions to make config larger
        }
    ]
    result = json.loads(sheetio.excel_extract(files, config, 8))

    assert len(result) == 50


# ---------------------------------------------------------------------------
# Combined stress test (10 files x 5 sheets x 500 rows x 10 columns)
# ---------------------------------------------------------------------------
@pytest.mark.slow
def test_combined_stress(tmp_path):
    """10 files, 5 sheets each (2 skipped), 500 rows, 10 columns, 4 workers."""
    files = []
    for i in range(10):
        f = _create_large_multirow_file(tmp_path, f"stress_{i}.xlsx", num_rows=500, num_columns=10, num_sheets=5)
        files.append(f)

    config = [
        {
            "sheets": ["Sheet*"],
            "skip_sheets": ["Sheet4", "Sheet5"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "label": "data",
                    "instructions": {
                        "row_range": [2, 501],
                        "unique_id": "A",
                        "stop_if_empty": "A",
                        "columns": {f"c{i}": get_column_letter(i + 1) for i in range(10)},
                    },
                }
            ],
        }
    ]
    result = json.loads(sheetio.excel_extract(files, config, 4))

    assert len(result) == 10

    total_rows = 0
    for file_data in result.values():
        for sheet_key, sheet_data in file_data.items():
            if sheet_key != "filepath" and "data" in sheet_data:
                total_rows += len(sheet_data["data"])

    # 10 files x 3 sheets (5 - 2 skipped) x 500 rows = 15000
    assert total_rows == 15000
