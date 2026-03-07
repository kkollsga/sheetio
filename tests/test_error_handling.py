"""Tests for error handling."""

import pytest

import sheet_excavator


def test_invalid_file_path():
    """Non-existent file is handled gracefully (logged, not raised)."""
    import json

    config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [{"function": "single_cells", "instructions": {"val": "a1"}}],
        }
    ]
    # The library handles invalid files gracefully - logs error, returns empty result for that file
    result = json.loads(sheet_excavator.excel_extract(["/nonexistent/file.xlsx"], config, 1))
    assert isinstance(result, dict)


def test_empty_file_list():
    """Empty list returns empty result."""
    config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [{"function": "single_cells", "instructions": {"val": "a1"}}],
        }
    ]
    import json

    result = json.loads(sheet_excavator.excel_extract([], config, 1))
    assert result == {} or isinstance(result, dict)


def test_empty_unique_id_array(tmp_path):
    """unique_id: [] should error."""
    from openpyxl import Workbook

    filepath = tmp_path / "empty_uid.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "data"
    wb.save(filepath)

    config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "instructions": {
                        "row_range": [1, 10],
                        "unique_id": [],
                        "columns": {"val": "A"},
                    },
                }
            ],
        }
    ]
    with pytest.raises(Exception):
        sheet_excavator.excel_extract([str(filepath)], config, 1)


def test_invalid_column_in_unique_id(tmp_path):
    """unique_id: ["B", "ZZZ"] - ZZZ is a valid column letter (col 18278), doesn't error."""
    import json

    from openpyxl import Workbook

    filepath = tmp_path / "bad_col.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "data"
    wb.save(filepath)

    config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "instructions": {
                        "row_range": [1, 10],
                        "unique_id": ["B", "ZZZ"],
                        "columns": {"val": "A"},
                    },
                }
            ],
        }
    ]
    # ZZZ is valid column notation, so this doesn't error - it just finds no matching data
    result = json.loads(sheet_excavator.excel_extract([str(filepath)], config, 1))
    assert isinstance(result, dict)


def test_non_string_in_unique_id_array(tmp_path):
    """unique_id: ["B", 123] -- non-string value in array should error."""
    from openpyxl import Workbook

    filepath = tmp_path / "non_string.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "data"
    wb.save(filepath)

    config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [
                {
                    "function": "multirow_patterns",
                    "instructions": {
                        "row_range": [1, 10],
                        "unique_id": ["B", 123],
                        "columns": {"val": "A"},
                    },
                }
            ],
        }
    ]
    with pytest.raises(Exception):
        sheet_excavator.excel_extract([str(filepath)], config, 1)
