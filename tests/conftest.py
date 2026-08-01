"""Shared fixtures for sheetio tests."""

import json
from datetime import datetime

import pytest
import sheetio
import xlwt
from openpyxl import Workbook


def run_extract(files, config, workers=1):
    """Run extraction and return parsed dict."""
    return json.loads(sheetio.excel_extract(files, config, workers))


def write_blank_string(sheet, row, col):
    """Write a genuine empty-string cell to a legacy .xls sheet.

    xlwt's public ``write("")`` silently downgrades to a BLANK record, which is a
    different thing on the wire. ``set_cell_text`` emits a real string record, which
    is what a cell holding "" looks like in files produced by Excel itself.
    """
    sheet.row(row).set_cell_text(col, "")


@pytest.fixture
def tmp_excel_simple(tmp_path):
    """Single sheet with mixed types: str, int, float, date, datetime."""
    filepath = tmp_path / "simple.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Title Value"
    ws["B2"] = "Description Text"
    ws["C3"] = 12345
    ws["D4"] = datetime(2024, 6, 15)
    ws["E5"] = datetime(2024, 6, 15, 14, 30, 0)
    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def tmp_excel_with_gaps(tmp_path):
    """8 rows with nulls at rows 4, 6-8 for stop_if_empty testing."""
    filepath = tmp_path / "gaps.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    data = [
        ("001", "Alice", 100, "A"),
        ("002", "Bob", 200, "B"),
        ("003", "Charlie", 300, None),
        (None, None, None, None),
        ("005", "Eve", 500, "E"),
        (None, None, None, None),
        (None, None, None, None),
        (None, None, None, None),
    ]
    for row_idx, (id_val, name, value, notes) in enumerate(data, start=1):
        if id_val is not None:
            ws.cell(row=row_idx, column=1, value=id_val)
        if name is not None:
            ws.cell(row=row_idx, column=2, value=name)
        if value is not None:
            ws.cell(row=row_idx, column=3, value=value)
        if notes is not None:
            ws.cell(row=row_idx, column=4, value=notes)
    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def tmp_excel_multisheet(tmp_path):
    """Sheets: School_A, School_B, School_C, Other_Sheet."""
    filepath = tmp_path / "multisheet.xlsx"
    wb = Workbook()
    for i, name in enumerate(["School_A", "School_B", "School_C", "Other_Sheet"]):
        if i == 0:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(name)
        ws["A1"] = f"Data from {name}"
        ws["B1"] = i * 100
    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def tmp_excel_dataframe(tmp_path):
    """Multi-row headers (rows 2-4) + data (rows 5-9)."""
    filepath = tmp_path / "dataframe.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "School_A"

    # Multi-row header
    for col, val in enumerate(["Student", "Math", "Science", "English", "Total"], start=2):
        ws.cell(row=2, column=col, value=val)
    for col, val in enumerate(["Information", "Score", "Score", "Score", "Score"], start=2):
        ws.cell(row=3, column=col, value=val)
    for col, val in enumerate(["Name", "(100)", "(100)", "(100)", "(300)"], start=2):
        ws.cell(row=4, column=col, value=val)

    # Data rows
    students = [
        ("Alice", 95, 88, 92, 275),
        ("Bob", 78, 85, 80, 243),
        ("Charlie", 88, 92, 85, 265),
        ("Diana", 92, 90, 95, 277),
        ("Eve", 85, 78, 88, 251),
    ]
    for row_idx, (name, math, sci, eng, total) in enumerate(students, start=5):
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=math)
        ws.cell(row=row_idx, column=4, value=sci)
        ws.cell(row=row_idx, column=5, value=eng)
        ws.cell(row=row_idx, column=6, value=total)

    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def tmp_excel_duplicates(tmp_path):
    """100 rows all with unique_id 'DUPLICATE_KEY'."""
    filepath = tmp_path / "duplicates.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Duplicates"
    for row in range(1, 101):
        ws.cell(row=row, column=1, value="DUPLICATE_KEY")
        ws.cell(row=row, column=2, value=row)
        ws.cell(row=row, column=3, value=f"Data_{row}")
    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def tmp_xls_simple(tmp_path):
    """Legacy .xls twin of tmp_excel_simple: str, int, float, date, datetime."""
    filepath = tmp_path / "simple.xls"
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")
    ws.write(0, 0, "Title Value")
    ws.write(1, 1, "Description Text")
    ws.write(2, 2, 12345)
    date_style = xlwt.easyxf(num_format_str="YYYY-MM-DD")
    datetime_style = xlwt.easyxf(num_format_str="YYYY-MM-DD HH:MM:SS")
    ws.write(3, 3, datetime(2024, 6, 15), date_style)
    ws.write(4, 4, datetime(2024, 6, 15, 14, 30, 0), datetime_style)
    wb.save(str(filepath))
    return str(filepath)


@pytest.fixture
def tmp_xls_blank_kinds(tmp_path):
    """Legacy .xls distinguishing the three ways a cell can look empty.

    Row 1 holds real values. Row 2 pairs a genuine empty-string cell (A2) with a
    real value, row 3 pairs a whitespace-only cell (A3) with a real value, and
    row 4 is never written at all.
    """
    filepath = tmp_path / "blank_kinds.xls"
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")
    ws.write(0, 0, "A1val")
    ws.write(0, 1, "B1val")
    write_blank_string(ws, 1, 0)
    ws.write(1, 1, "B2val")
    ws.write(2, 0, "   ")
    ws.write(2, 1, "B3val")
    # Row 4 (index 3) deliberately left unwritten.
    wb.save(str(filepath))
    return str(filepath)


@pytest.fixture
def tmp_xls_empty_string_row(tmp_path):
    """Legacy .xls where the separator row holds empty strings, not blanks.

    Rows 1-2 are data, row 3 is empty strings in every extracted column, row 4 is
    data again. A stop condition must halt at row 3 regardless of whether the file
    stores that row as blanks or as empty strings.
    """
    filepath = tmp_path / "empty_string_row.xls"
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Sheet1")
    for row_idx, (id_val, name) in enumerate([("001", "Alice"), ("002", "Bob")]):
        ws.write(row_idx, 0, id_val)
        ws.write(row_idx, 1, name)
    write_blank_string(ws, 2, 0)
    write_blank_string(ws, 2, 1)
    ws.write(3, 0, "004")
    ws.write(3, 1, "Dave")
    wb.save(str(filepath))
    return str(filepath)


@pytest.fixture
def tmp_excel_whitespace_row(tmp_path):
    """.xlsx twin of tmp_xls_empty_string_row using whitespace-only cells."""
    filepath = tmp_path / "whitespace_row.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"], ws["B1"] = "001", "Alice"
    ws["A2"], ws["B2"] = "002", "Bob"
    ws["A3"], ws["B3"] = "   ", "   "
    ws["A4"], ws["B4"] = "004", "Dave"
    wb.save(filepath)
    return str(filepath)


@pytest.fixture
def tmp_xls_multirow(tmp_path):
    """Legacy .xls with 4 data rows for multirow/dataframe extraction."""
    filepath = tmp_path / "multirow.xls"
    wb = xlwt.Workbook()
    ws = wb.add_sheet("TestSheet")
    header = ["ID", "Name", "Value"]
    for col, val in enumerate(header):
        ws.write(0, col, val)
    rows = [("001", "Alice", 100), ("002", "Bob", 200), ("003", "Charlie", 300)]
    for row_idx, (id_val, name, value) in enumerate(rows, start=1):
        ws.write(row_idx, 0, id_val)
        ws.write(row_idx, 1, name)
        ws.write(row_idx, 2, value)
    wb.save(str(filepath))
    return str(filepath)


@pytest.fixture
def tmp_excel_composite(tmp_path):
    """Project+Year data for composite key testing."""
    filepath = tmp_path / "composite.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    data = [
        ("ProjectA", 2024, 100000, "Active"),
        ("ProjectA", 2025, 120000, "Planned"),
        ("ProjectB", 2024, 80000, "Active"),
        ("ProjectB", 2025, 90000, "Planned"),
    ]
    for row_idx, (project, year, budget, status) in enumerate(data, start=1):
        ws.cell(row=row_idx, column=2, value=project)
        ws.cell(row=row_idx, column=3, value=year)
        ws.cell(row=row_idx, column=4, value=budget)
        ws.cell(row=row_idx, column=5, value=status)
    wb.save(filepath)
    return str(filepath)
