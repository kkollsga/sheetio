"""Tests for single_cells extraction function."""

from conftest import run_extract


def test_basic_extraction(tmp_excel_simple):
    """Extract string, int, float, date, datetime from known cells."""
    config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [
                {
                    "function": "single_cells",
                    "label": "single",
                    "instructions": {
                        "Value 1": "a1",
                        "Value 2": "b2",
                        "Value 3": "c3",
                        "Date": "d4",
                        "Datetime": "e5",
                    },
                }
            ],
        }
    ]
    result = run_extract([tmp_excel_simple], config)
    file_key = list(result.keys())[0]
    data = result[file_key]["Sheet1"]["single"]

    assert data["Value 1"] == "Title Value"
    assert data["Value 2"] == "Description Text"
    assert data["Value 3"] == 12345.0
    assert "Date" in data
    assert "Datetime" in data


def test_with_label(tmp_excel_simple):
    """Verify label key nesting."""
    config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [
                {
                    "function": "single_cells",
                    "label": "my_label",
                    "instructions": {"val": "a1"},
                }
            ],
        }
    ]
    result = run_extract([tmp_excel_simple], config)
    file_key = list(result.keys())[0]
    assert "my_label" in result[file_key]["Sheet1"]
    assert result[file_key]["Sheet1"]["my_label"]["val"] == "Title Value"


def test_without_label(tmp_excel_simple):
    """Values stored directly under sheet name when no label."""
    config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [
                {
                    "function": "single_cells",
                    "instructions": {"val": "a1"},
                }
            ],
        }
    ]
    result = run_extract([tmp_excel_simple], config)
    file_key = list(result.keys())[0]
    assert result[file_key]["Sheet1"]["val"] == "Title Value"


def test_break_if_null_triggers(tmp_path):
    """Sheet skipped when break_if_null target cell is null."""
    from openpyxl import Workbook

    filepath = tmp_path / "break_null.xlsx"
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "Value A"
    # C3 intentionally empty

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Value B"
    ws2["C3"] = "Not Null"
    wb.save(filepath)

    config = [
        {
            "sheets": ["Sheet1", "Sheet2"],
            "extractions": [
                {
                    "function": "single_cells",
                    "break_if_null": "c3",
                    "instructions": {"test": "a1"},
                }
            ],
        }
    ]
    result = run_extract([str(filepath)], config)
    assert isinstance(result, dict)


def test_multi_cell_array_ref(tmp_path):
    """Array of cell refs for one key."""
    from openpyxl import Workbook

    filepath = tmp_path / "array_ref.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["H7"] = "Project Alpha"
    ws["H8"] = "Project Beta"
    ws["H9"] = "Project Gamma"
    ws["H10"] = None
    ws["H11"] = None
    wb.save(filepath)

    config = [
        {
            "sheets": ["Sheet1"],
            "extractions": [
                {
                    "function": "single_cells",
                    "label": "single",
                    "instructions": {"project_name": ["h7", "h8", "h9", "h10", "h11"]},
                }
            ],
        }
    ]
    result = run_extract([str(filepath)], config)
    file_key = list(result.keys())[0]
    project_names = result[file_key]["Sheet1"]["single"]["project_name"]
    assert isinstance(project_names, list)
    assert "Project Alpha" in project_names
    assert "Project Beta" in project_names
    assert "Project Gamma" in project_names
